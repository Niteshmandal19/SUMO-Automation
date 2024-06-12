import os
import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom
import subprocess

def generate_nodes_xml(nodes_data, output_file):
    root = ET.Element("nodes")
    for _, row in nodes_data.iterrows():
        node = ET.SubElement(root, "node")
        node.set("id", str(row["node_id"]))
        node.set("x", str(row["x"]))
        node.set("y", str(row["y"]))
        if pd.notnull(row["z"]):
            node.set("z", str(row["z"]))
        if not pd.isna(row["node_type"]):
            node.set("type", row["node_type"])

    tree = ET.ElementTree(root)
    tree.write(output_file, xml_declaration=True, encoding="utf-8")

def xlsx_to_edge_xml(xlsx_file, xml_file, sheet_name="Edges"):
    wb = openpyxl.load_workbook(xlsx_file)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    sheet = wb[sheet_name]

    headers = [cell.value for cell in sheet[1]]

    edges = ET.Element('edges')

    for row in sheet.iter_rows(min_row=2, values_only=True):
        edge_data = dict(zip(headers, row))

        edge = ET.SubElement(edges, 'edge')

        edge.set('id', str(edge_data.get('Edge ID', '')))

        from_node = str(edge_data.get('From Node', ''))
        to_node = str(edge_data.get('To Node', ''))
        edge.set('from', from_node)
        edge.set('to', to_node)

        optional_attrs = [
            'Type', 'numLanes', 'Speed (m/s)', 'Priority', 'Length (m)',
            'Shape', 'SpreadType', 'Allow', 'Disallow', 'Width (m)', 'Name',
            'endOffset (m)', 'sidewalkWidth (m)', 'bikeLaneWidth (m)', 'distance (m)'
        ]

        attr_mapping = {
            'Type': 'type',
            'numLanes': 'numLanes',
            'Speed (m/s)': 'speed',
            'Priority': 'priority',
            'Length (m)': 'length',
            'Shape': 'shape',
            'SpreadType': 'spreadType',
            'Allow': 'allow',
            'Disallow': 'disallow',
            'Width (m)': 'width',
            'Name': 'name',
            'endOffset (m)': 'endOffset',
            'sidewalkWidth (m)': 'sidewalkWidth',
            'bikeLaneWidth (m)': 'bikeLaneWidth',
            'distance (m)': 'distance'
        }

        for xlsx_attr, xml_attr in attr_mapping.items():
            if xlsx_attr in edge_data and edge_data[xlsx_attr] is not None:
                edge.set(xml_attr, str(edge_data[xlsx_attr]))

    rough_string = ET.tostring(edges, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml = reparsed.toprettyxml(indent="  ")

    with open(xml_file, "w") as f:
        f.write(pretty_xml)

if __name__ == "__main__":
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Full path to the Excel file
        file_path = os.path.join(script_dir, "data.xlsx")

        # Print available sheet names to debug
        with pd.ExcelFile(file_path, engine='openpyxl') as xls:
            print("Available sheets:", xls.sheet_names)

        # Read data from Excel file using openpyxl engine
        nodes_data = pd.read_excel(file_path, sheet_name="Nodes", engine='openpyxl')
        print("Data read successfully.")

        # Generate nodes XML file
        generate_nodes_xml(nodes_data, os.path.join(script_dir, "nodes.nod.xml"))
        print("Nodes XML file generated successfully.")

        # Generate edges XML file
        xlsx_to_edge_xml(file_path, os.path.join(script_dir, "edges.edg.xml"))
        print("Edges XML file generated successfully.")

        # Full paths to SUMO tools (adjust these paths if necessary)
        netconvert_path = "D:/Softwares/Sumo/bin/netconvert.exe"
        random_trips_path = "D:/Softwares/Sumo/tools/randomTrips.py"
        duarouter_path = "D:/Softwares/Sumo/bin/duarouter.exe"

        # Print paths to debug
        print(f"netconvert_path: {netconvert_path}")
        print(f"random_trips_path: {random_trips_path}")
        print(f"duarouter_path: {duarouter_path}")

        # Check if SUMO executables exist
        if not os.path.exists(netconvert_path):
            raise FileNotFoundError(f"'{netconvert_path}' not found.")
        if not os.path.exists(random_trips_path):
            raise FileNotFoundError(f"'{random_trips_path}' not found.")
        if not os.path.exists(duarouter_path):
            raise FileNotFoundError(f"'{duarouter_path}' not found.")

        # Step 3: Create sumotest.net.xml
        subprocess.run([netconvert_path, "--node-files=nodes.nod.xml", "--edge-files=edges.edg.xml", "--output-file=sumotest.net.xml"], check=True)
        print("sumotest.net.xml created successfully.")

        # Step 4: Create trips
        subprocess.run(["python", random_trips_path, "-n", "sumotest.net.xml", "-e", "1000", "-o", "sumotest.trips.xml"], check=True)
        print("sumotest.trips.xml created successfully.")

        # Step 5: Create routes file
        subprocess.run([duarouter_path, "-n", "sumotest.net.xml", "--route-files", "sumotest.trips.xml", "-o", "sumotest.rout.xml", "--ignore-errors"], check=True)
        print("sumotest.rout.xml created successfully.")

        # Step 6: Create sumotest.sumo.cfg
        with open(os.path.join(script_dir, "sumotest.sumo.cfg"), "w") as f:
            f.write("""<?xml version="1.0" encoding="UTF-8"?>
<configuration xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://sumo.dlr.de/xsd/sumoConfiguration.xsd">
    <input>
        <net-file value="sumotest.net.xml"/>
        <route-files value="sumotest.trips.xml"/>
    </input>
    <time>
        <begin value="0"/>
        <end value="1000"/>
    </time>
</configuration>""")
        print("sumotest.sumo.cfg created successfully.")

    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except ValueError as e:
        print(f"Value error: {e}")
    except subprocess.CalledProcessError as e:
        print(f"Subprocess error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
